#!/usr/bin/env python3

'''
    File:       jenkins2xl.py
    Copyright:  (c) 2016 by VMware, Inc.  All right reserved

    This script takes one or more log files containing test pass/fail data
    and summarizes the into a native Excel file (.xlsx).  There is lots of
    embedded assumptions in this, so please read below.

    LOG FILES
    ----------
    The log files are usually named 'log' but no longer matters to this script.
    The path name is no longer important as the test name is in the log file.

    LOG FILE CONTENT
    ----------------
    There are 3 things that the script is looking for inside the log file and
    changes to the log file format may affect the script's functionality.

        1) the test name (see more information on test names below)
        2) the names of the failed tests
        3) the pass/fail status of this test

    TEST NAMES: A test name is of the form:
        'esxui-<testname>-<os>-<browser>'
    where:
        testname is the name of the test
        os       is the name of the os running the browser for the tests
        browser  is the name of the browser used for the test
    NOTE: the test name may contain dashes ('-') but the others may not.
    NOTE: It is OK to change the data structures in the XLSX class to add
          or remove tests, OS's, or Browsers.  The format of the Excel
          tabs will change only for tabs on that test date forward.
    The os must match one of the names in XLSX.OS_NAMES array.
    The browser must match one of the names in XLSX.BROWSER_NAMES array.
    The testname must match one of the names in XLSX._NAMES array.
    The definition of 'match' is defined by XLSX.add.add_helper()

    The script will search for a line matching this regular expression:
        '^\[esxui(-\w+){3,}\]'
    Definition:
        - starting at the beginning of a line
        - match an open-square-bracket character '['
        - then match the word 'esxui'
        - then 3 to infinity times:
            match a single dash '-' followed by [a-zA-Z0-9_] 1 or more times
        - match a close-square-bracket character ']'

    LOG FAILURES and PENDING TESTS:  The log optionally will have lines like:
        Failures:
        1) VM Hot Edit - add hard disk test cases Hot Edit VM -  Add new hard disk - Shares High
    and
        Pending:
        1) VM Hot Edit - add other device test cases Hot Edit - Add USB

    It will collect the numbered lines and add them to the log.  These can come
    in any order and the script will handle that.


    PASS/FAIL STATUS: The log will have two consecutive lines that look like:

        46 specs, 0 failures, 3 pending specs
        Finished in 123.456 seconds

    We search each file for the line beginning with 'Finished' and then look
    at the preceeding line.  There may be no other lines in the log that begin
    with 'Finished' for this to work correctly currently.  Then we back up one
    line and parse the numbers before the word 'specs' and 'failures'.  These
    are the number of test cases and the number of test failures.

    The modification date of the log file is used to create the tabs in the
    Excel file.  The current format of the tab names are yyyy-mm-dd.  The data
    from a log file created on a given day will be added to the that tab.

    WARNING: if two tests are run on the same day over the same tests, the
    last process log file will OVERWRITE the previous data.

    XLSX CLASS
    ----------
    The XLSX class creates a new .xlsx file if the given file name does not exist.
    It creates a new tab in the workbook the first time is sees a date, which
    implies that changes to the class will not be reflected in a tab's format
    until it sees data for a new day.

    The XLSX Python Class is meant to be as data driven as possible, making as
    few assumptions as possible.  There are a few assumptions however:
        1) The column for the test ID is column 'A' (hardcoded)
        2) The column for the test Name is column 'B' (hardcoded)
        3) The column for the test data will start in column 'C' (hardcoded)
        4) There is always a single 'gap' column between browsers
        5) There is always a single 'gap' column between os's
        6) The values in the percentage fields must be numbers from 0.0 - 1.0
        7) The red-yellow-green stop light icons change based on numbers, not percentages
        8) The Pass % column data is formatted to be a number displayed as percent

    Provided you do these before any tests are added for a given day, it is OK to:
        - Test Names: add/remove a name or change the order/spelling
        - Browser: add/remove a name or change the order/spelling (see LAYOUT structure)
        - OS: add/remove a name or change the order/spelling (see LAYOUT structure)
        - TEST RESULT CLASSES - change the spelling, but not the order
        - change any color (can be changed anytime, but only reflected in a new tab)

    OTHER CONSIDERATIONS
    ---------------------
    The Python package 'openpyxl' must be installed for this script to work.  The
    import of the package is protected with a try/except block to remind users.
'''


# ------------------------------------------------------------------------------
# IMPORTED LIBRARIES
#
import argparse
import datetime
import logging
import os
from random import randint
import re
import string
import sys
import time
import xml.etree.ElementTree as ET

try:
    from openpyxl.formatting.rule import FormatObject, IconSet, Rule
    from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
    from openpyxl.styles.fills import FILL_SOLID
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl import Workbook, load_workbook
except:
    print("You must install the openpyxl Python library to use this script.",
          file=sys.stderr)
    print("Try the command:", file=sys.stderr)
    print("   pip install openpyxl", file=sys.stderr)
    print("For more information: http://openpyxl.readthedocs.io/en/default/",
          file=sys.stderr)
    exit(1)

# --- CONSTANTS ---
_VERSION = '1.0'

# --- KEY CONSTANTS ---
OS = 'OS'
BROWSER = 'BROWSER'
RESULT = 'RESULT'
GRAND = 'GRAND'

# --- EXIT CODEs ---
_EXIT_SUCCESS = 0
_EXIT_ERROR = 1
_EXIT_PROGRAM_ERROR = 2
_EXIT_CONTROL_C_EXCEPTION = 3

# --- LOG FILE REGULAR EXPRESSIONS ---
RE_TEST_NAME = re.compile('esxui(-\w+){3,}')
RE_RESULTS = re.compile('(\d+) spec[s]?, (\d+) failure[s]?(, (\d+) pending spec[s]?)?')
RE_FINISHED = re.compile('Finished in [\d.]+ seconds')
RE_TEST_FAIL_START = re.compile('Failures:$')
RE_TEST_PEND_START = re.compile('Pending:$')
RE_TEST_DATA_LINE = re.compile('\d+\)\s+(.*)')
RE_BROWSER_CHROME = re.compile('(chrome)', re.IGNORECASE)
RE_BROWSER_FIREFOX = re.compile('(firefox)', re.IGNORECASE)
RE_BROWSER_IE = re.compile('(iexplorer)', re.IGNORECASE)
RE_BROWSER_SAFARI = re.compile('(safari)', re.IGNORECASE)
RE_VIC_FLEX_TEST_FAIL_START = re.compile('FAILED: execute')
RE_VIC_FLEX_TEST_PASS_START = re.compile('PASSED: execute')
RE_VIC_H5_TEST_FAIL_START = re.compile('.*Vannyo reports \'FAILED\' for test \'(.*)\'')
RE_VIC_H5_TEST_PASS_START = re.compile('.*Vannyo reports \'PASSED\' for test \'(.*)\'')

# ------------------------------------------------------------------------------
class XLSX(object):

    # ---- OS NAMES ----
    OS_UBUNTU = "Ubuntu"
    OS_WINDOWS = "Windows"
    OS_MACOSX = "Mac"
    OS_BLANK = ""
    OS_NAMES = [
        OS_UBUNTU,
        OS_WINDOWS,
        OS_MACOSX,
        OS_BLANK
    ]

    # ---- BROWSER NAMES ----
    BROWSER_CHROME = "Chrome"
    BROWSER_FIREFOX = "Firefox"
    BROWSER_IE = "IE11"
    BROWSER_SAFARI = "Safari"
    BROWSER_NAMES = (
        BROWSER_CHROME,
        BROWSER_FIREFOX,
        BROWSER_IE,
        BROWSER_SAFARI
    )

    # ---- OS/BROWSER LAYOUT ----
    BROWSER_SUBTOTAL = 'Test Subtotals'
    LAYOUT = [
        {OS_UBUNTU:     [BROWSER_CHROME, BROWSER_FIREFOX]},
        {OS_WINDOWS:    [BROWSER_CHROME, BROWSER_FIREFOX, BROWSER_IE]},
        {OS_MACOSX:     [BROWSER_CHROME, BROWSER_FIREFOX, BROWSER_SAFARI]},
        {OS_BLANK:      [BROWSER_SUBTOTAL]}     # Hack for subtotals
    ]

    # ---- TEST RESULT CLASSES ----
    RESULT_TOTAL = "Total"
    RESULT_FAIL = "Fail"
    RESULT_PCT = "% Pass"
    RESULT_NAMES = (
        RESULT_TOTAL,
        RESULT_FAIL,
        RESULT_PCT
    )

    # ---- TEST NAMES ----
    TEST_NAMES = [
        "18-1-VIC-UI-Installer-60",
        "18-1-VIC-UI-Installer-65",
        "18-2-VIC-UI-Uninstaller-60",
        "18-2-VIC-UI-Uninstaller-65",
        "18-3-VIC-UI-Upgrader-60",
        "18-3-VIC-UI-Upgrader-65",
        "18-4-VIC-UI-Plugin-tests-60",
        "18-4-VIC-UI-Plugin-tests-65"
    ]

    GRAYOUT_LAYOUT = [
        {OS_UBUNTU: {
            BROWSER_CHROME: [],
            BROWSER_FIREFOX: TEST_NAMES[0:6]
        }},
        {OS_WINDOWS: {
            BROWSER_CHROME: TEST_NAMES[0:6],
            BROWSER_FIREFOX: TEST_NAMES[0:6],
            BROWSER_IE: TEST_NAMES[0:6]
        }},
        {OS_MACOSX: {
            BROWSER_CHROME: [],
            BROWSER_FIREFOX: TEST_NAMES[0:6],
            BROWSER_SAFARI: TEST_NAMES[0:6]
        }}
    ]

    # ---- SUBTOTAL ROW NAMES ----
    SUBTOTAL_NAMES = [
        "Browser Totals:",
        "OS Totals:",
        "Grand Total:"
    ]

    # ---- BORDER FORMATTING ----
    THIN = Side(border_style="thin", color="000000")
    THICK = Side(border_style="thick", color="000000")
    THIN_BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
    THICK_BORDER = Border(top=THICK, left=THICK, right=THICK, bottom=THICK)

    # ---- GAP FORMATTING (OK to resize) ----
    WIDTH_ID = 3.0
    WIDTH_OS_GAP = 1.0

    # ---- FIELD WIDTH FORMATTING (NOT OK to make smaller) ----
    WIDTH_TEST = len(max(TEST_NAMES, key=len))
    WIDTH_RESULT = len(max(RESULT_NAMES, key=len))

    # ---- VERTICAL PLACEMENT CONSTANTS (OK to change) ----
    ROW_OS = 1
    ROW_BROWSER = ROW_OS + 1
    ROW_TITLES = ROW_OS + 2
    ROW_TESTS = ROW_OS + 3

    # ---- COLORS: NAMED (OK to change) ----
    COLOR_WHITE = 0xFFFFFF
    COLOR_BLACK = 0x000000
    COLOR_GREEN = 0x00FF00
    COLOR_YELLOW = 0xFFFF00
    COLOR_RED = 0xFF0000
    COLOR_DKGRAY = 0x404040
    COLOR_OS_UBUNTU_ORANGE = 0xE95420   # Ubuntu Orange
    COLOR_OS_WINDOWS_BLUE = 0x396DA6    # Microsoft Blue
    COLOR_OS_MAC_SILVER = 0xE0E0E0      # Apple Silver(-ish)
    COLOR_ID_FONT = COLOR_WHITE
    COLOR_ID_BKGD = COLOR_DKGRAY
    COLOR_TEST_FONT = COLOR_WHITE
    COLOR_TEST_BKGD = COLOR_DKGRAY
    COLOR_BROWSER_CHROME = 0x4285F4
    COLOR_BROWSER_FIREFOX = 0xD37D2B
    COLOR_BROWSER_IE = 0x0E4678
    COLOR_BROWSER_SAFARI = 0x0E8596

    # ---- COLORS: OS HEADINGS (OK to change) ----
    COLOR_OS = {
        OS_UBUNTU:  COLOR_OS_UBUNTU_ORANGE,
        OS_WINDOWS: COLOR_OS_WINDOWS_BLUE,
        OS_MACOSX:  COLOR_OS_MAC_SILVER,
        OS_BLANK:   COLOR_WHITE
    }

    # ---- COLORS: BROWSER HEADINGS (OK to change) ----
    # Ref: http://paletton.com/#uid=73D0u0kllllaFw0g0qFqFg0w0aF
    COLOR_BROWSER = {
        BROWSER_CHROME:     {'fill': COLOR_BROWSER_CHROME,
                             'font': COLOR_WHITE},
        BROWSER_FIREFOX:    {'fill': COLOR_BROWSER_FIREFOX,
                             'font': COLOR_WHITE},
        BROWSER_IE:         {'fill': COLOR_BROWSER_IE,
                             'font': COLOR_WHITE},
        BROWSER_SAFARI:     {'fill': COLOR_BROWSER_SAFARI,
                             'font': COLOR_WHITE},
        BROWSER_SUBTOTAL:   {'fill': COLOR_BLACK, 'font': COLOR_WHITE},
    }

    # ---- COLORS: RESULT HEADINGS (OK to change, but please don't) ----
    COLOR_RESULT = {
        RESULT_TOTAL:   COLOR_WHITE,
        RESULT_FAIL:    COLOR_RED,
        RESULT_PCT:     COLOR_GREEN
    }

    # ---- COLORS: SUBTOTALs (OK to change) ----
    COLOR_SUBTOTAL = {
        BROWSER:  {'fill': 0x8b4513, 'font': COLOR_WHITE},
        OS:       {'fill': 0xcd853f, 'font': COLOR_WHITE},
        GRAND:    {'fill': 0xdeb887, 'font': COLOR_WHITE}
    }

    # ---- COLORS: TEST THRESHOLDS (OK to change the Green and Yellow values) ----
    # WARNING: THIS WAS TRICKY TO GET RIGHT - CHANGE AT YOUR OWN RISK
    THRESHOLD_GREEN = 0.97     # Test run is green if value >= this
    THRESHOLD_YELLOW = 0.90    # Test run is yellow  if value >= this
    THRESHOLD_RED = 0.0      # Test run is red if value >= this

    COLOR_VALUE_THRESHOLDS = {
        THRESHOLD_GREEN:    COLOR_GREEN,
        THRESHOLD_YELLOW:   COLOR_YELLOW,
        THRESHOLD_RED:      COLOR_RED
    }

    RFO1 = FormatObject(type='num', val=THRESHOLD_RED)
    RFO2 = FormatObject(type='num', val=THRESHOLD_YELLOW)
    RFO3 = FormatObject(type='num', val=THRESHOLD_GREEN)
    RESULT_IS = IconSet(iconSet='3TrafficLights1',
                        cfvo=[RFO1, RFO2, RFO3],
                        showValue=None,
                        percent=None,
                        reverse=None)
    RESULT_RULE = Rule(type='iconSet', iconSet=RESULT_IS)
    RESULT_FORMAT = '0.0%'

    RESULT_FORMULA = '''
=IF(
    OR(
        ISBLANK(INDIRECT("C[-2]",FALSE)),
        INDIRECT("C[-2]",FALSE)=0
    ),
    "",
    (1-(INDIRECT("C[-1]",FALSE)/INDIRECT("C[-2]",FALSE)))
)
    '''.strip().replace('\n', '').replace(' ', '').replace(',', ', ')

    # --------------------------------------------------------------------
    # -- METHODSx
    # --------------------------------------------------------------------
    def __init__(self, wbName='wb_'+str(int(time.time()))+'.xlsx'):
        self.__wbName = wbName
        self.__wb = None
        self.create()

    # --------------------------------------------------------------------
    @property
    def wbName(self):
        return self.__wbName

    # --------------------------------------------------------------------
    @property
    def wb(self):
        return self.__wb

    # --------------------------------------------------------------------
    @staticmethod
    def cs(rgbVal):
        ''' Returns an 8-character hex color value for use with openpyxl '''
        return "{:08X}".format(rgbVal)

    # --------------------------------------------------------------------
    @staticmethod
    def color(rgbVal):
        ''' Returns an 8-character hex color value for use with openpyxl '''
        return Color(rgb=XLSX.cs(rgbVal))

    # --------------------------------------------------------------------
    @staticmethod
    def cc2N(str_col):
        ''' Returns the column letter (or letters) to an index '''
        return column_index_from_string(str_col) - 1

    # --------------------------------------------------------------------
    @staticmethod
    def cn2C(idx):
        N = len(string.ascii_uppercase)
        # Common fast case
        if idx >= 0 and idx < N:
            return string.ascii_uppercase[idx]
        return get_column_letter(idx+1)

    # --------------------------------------------------------------------
    @staticmethod
    def ccNext(str_col):
        ''' Returns the next column character for an Excel spreadsheet '''
        return XLSX.cn2C(XLSX.cc2N(str_col) + 1)

    # --------------------------------------------------------------------
    @staticmethod
    def ccPrev(str_col):
        ''' Returns the prev column character for an Excel spreadsheet '''
        return XLSX.cn2C(XLSX.cc2N(str_col) - 1)

    # --------------------------------------------------------------------
    def solidFill(self, rgbVal):
        ''' Returns openpyxl.styles.PatternFill '''
        c = self.color(rgbVal)
        return PatternFill(fill_type=FILL_SOLID, fgColor=c, bgColor=c)

    # --------------------------------------------------------------------
    def setActiveByID(self, idx=0):
        ''' Returns sets the active sheet by index '''
        assert idx >= 0 and idx < len(self.wb.sheetnames)
        self.wb.active = idx

    # --------------------------------------------------------------------
    def setActiveByName(self, name):
        ''' Returns sets the active sheet by name '''
        assert name in self.wb.sheetnames
        self.setActiveByID(self.wb.sheetnames.index(name))

    # --------------------------------------------------------------------
    def layout(self, layoutIdx):
        ''' Returns the OS name and browser list '''
        assert layoutIdx < len(self.LAYOUT), 'layout - index out of range'
        layoutDict = self.LAYOUT[layoutIdx]
        osName = list(layoutDict.keys())[0]
        return (osName, layoutDict[osName])

    # --------------------------------------------------------------------
    def browsersByOS(self, osName):
        ''' Returns the list of browsers for the given OS name '''
        return self.layout(self.OS_NAMES.index(osName))

    # --------------------------------------------------------------------
    @property
    def width(self):
        ''' Returns the width of the current layout '''
        cols = 0
        for i in range(len(self.LAYOUT)):
            (os, b) = self.layout(i)
            cols += len(b)              # Count of browsers for each OS
        blankCols = cols - 1            # Blank col between each browser set
        cols *= 3                       # Each browser comprises 3 columns
        return cols + blankCols + 2     # Two initial cols for ID and Name

    # --------------------------------------------------------------------
    def add(self, testDate, logfile):
        logging.debug("Adding test: %s %s %s %s %d %d" % (
                testDate,
                logfile.test_name_os,
                logfile.test_name_browser,
                logfile.test_name_suite,
                logfile.num_tests,
                logfile.num_fails
                )
        )

        def add_helper(val, valList, typeName):
            uVals = [_.upper() for _ in valList]
            uVal = val.upper()
            if uVal not in uVals:
                found = False
                for i, uval in enumerate(uVals):
                    if uVal.startswith(uval) or uVal.find(uval) >= 0:
                        val = valList[i]
                        found = True
                        break
                if not found:
                    _fmt = "\'%s\' is not a valid %s name: os:{} brwsr:{} test:{}".format(osName, browserName, testName)
                    raise ValueError(_fmt % (val, typeName))
            else:
                val = valList[uVals.index(uVal)]
            return val

        # Match the OS name
        orig = logfile.test_name_os
        osName = add_helper(logfile.test_name_os, self.OS_NAMES, OS)
        logging.debug("%s matched %s" % (orig, osName))

        # Match the Browser name
        orig = logfile.test_name_browser
        browserName = add_helper(logfile.test_name_browser, self.BROWSER_NAMES, 'browser')
        logging.debug("%s matched %s" % (orig, browserName))

        # Match the test name
        orig = logfile.test_name_suite
        testName = add_helper(logfile.test_name_suite, self.TEST_NAMES, 'test')
        logging.debug("%s matched %s" % (orig, testName))

        # Create our tab if necessary
        dt = str(testDate).split()[0]
        if dt not in self.wb.sheetnames:
            self.createTab(dt)
        self.setActiveByName(dt)

        # Find the place for the data
        row = self.ROW_TESTS + self.TEST_NAMES.index(testName)
        # print("Test row for %s: %s" % (testName, row))
        colMap = {}
        col = 2     # Zero-based column for this case
        for idx in range(len(self.LAYOUT)):
            (os, browsers) = self.layout(idx)
            colMap.setdefault(os, dict())
            for browser in browsers:
                colMap[os][browser] = col
                col += 4     # 3 data columns + 1 gap column
        col = colMap[osName][browserName]

        ws = self.wb.get_active_sheet()
        specCellID = "{}{}".format(self.cn2C(col), row)
        failCellID = "{}{}".format(self.cn2C(col+1), row)
        # print("Setting cell [%s] to %d " % (specCellID, specs))
        # print("Setting cell [%s] to %d " % (failCellID, fails))
        ws[specCellID] = int(logfile.num_tests)
        hyperlink_font = Font(name='Monaco', size=10, underline="single", color="FF0000FF")
        ws[specCellID].hyperlink = "%s" % (logfile.log_bundle_fname)
        ws[specCellID].font = hyperlink_font
        ws[failCellID] = int(logfile.num_fails)

        # ---------------------------------------------------------
        # Update the Failure list and Pending test list
        # The headings are in column A and data in column B
        # baseRow is one past the 3 subtotal rows
        # We collect up the existing data, add any new data and rewrite
        HDG_FAILED = 'Failed Tests'
        HDG_PENDING = 'Pending Tests'
        HDG_FILL = {HDG_FAILED: self.COLOR_RED, HDG_PENDING: self.COLOR_YELLOW}

        def collect_data(existing, hdg, row):
            existing.setdefault(hdg, dict())
            while ws['A'+str(row)].value:
                count = ws['A'+str(row)].value
                data = ws['B'+str(row)].value
                (cases, data) = data.split(':', 1)
                cases = cases.strip()
                data = data.strip()
                existing[hdg].setdefault(data, {'Count': 0, 'Cases': set()} )
                existing[hdg][data]['Count'] += int(count)
                for case in cases.split(','):
                    case = case.strip()
                    existing[hdg][data]['Cases'].add(case)
                row += 1
            return row

        def merge_data(existing, hdg, case, map):
            existing.setdefault(hdg, dict())
            for data in map:
                existing[hdg].setdefault(data, {'Count': 0, 'Cases': set()} )
                existing[hdg][data]['Count'] += 1
                existing[hdg][data]['Cases'].add(case)

        # Collect existing data from the worksheet
        currRow = baseRow = self.ROW_TITLES + len(self.TEST_NAMES) + 3 + 2 + 5
        existing = {}
        for i in range(2):      # Loop for Failures and Pending
            hdg = ws['B'+str(currRow)].value  # Get heading or blank
            if not hdg:     # T -> done
                break
            currRow = collect_data(existing, hdg, currRow+1) + 1

        # Clear out the old data
        for row in range(baseRow, currRow+1):
            for col in range(1, 3):
                coordinate = (row, col)
                if coordinate in ws._cells:
                    del ws._cells[coordinate]

        # Merge in data from log file
        case = logfile.test_name_os[0] + logfile.test_name_browser[0]
        merge_data(existing, HDG_FAILED, case, logfile.tests_failed_map)
        merge_data(existing, HDG_PENDING, case, logfile.tests_pending_map)

        # Write out the updated data
        fixedFont = Font(name='Monaco', size=10)
        lastCol = get_column_letter(self.width)
        currRow = baseRow
        for hdg in [HDG_FAILED, HDG_PENDING]:
            if hdg in existing and len(existing[hdg]):
                hdgCell = ws['B'+str(currRow)]
                hdgCell.value = hdg
                hdgCell.fill = self.solidFill(HDG_FILL[hdg])
                ws.merge_cells('B'+str(currRow)+":"+lastCol+str(currRow))
                for data in sorted(existing[hdg]):
                    currRow += 1
                    ws['A'+str(currRow)].value = existing[hdg][data]['Count']
                    cases = ', '.join(sorted(existing[hdg][data]['Cases']))
                    data = '{:30s}'.format(cases) + ': ' + data
                    ws['B'+str(currRow)].value = data
                    ws['B'+str(currRow)].font = fixedFont
                    ws.merge_cells('B'+str(currRow)+":AK"+str(currRow))
                currRow += 1


    # ------------
    def create(self):
        assert self.wb is None, "Workbook exists: {}".format(self.wbName)
        wbName = self.wbName
        if os.path.isfile(wbName) and os.access(wbName, os.W_OK):
            logging.info("\'%s\' exists and is writable" % (wbName))
            self.__wb = load_workbook(wbName)
            self.setActiveByID(len(self.wb.sheetnames) - 1)
        elif os.path.isfile(wbName):
            logging.error("File \'%s\' exists but is NOT writable" % (wbName))
            raise AttributeError("File \'%s\' exists but is NOT writable" %
                                 (wbName))
        else:
            logging.info("\'%s\' does not exist, creating..." % (wbName))
            self.__wb = Workbook()
            self.setActiveByID()
            self.save()


    # --------------------------------------------------------------------
    def save(self):
        assert self.wb is not None, "XLSX Class has no workbook name"
        self.wb.active = min(1, len(self.wb.sheetnames)-1)  # If 1 sheet
        self.wb.save(self.wbName)

    # --------------------------------------------------------------------
    def __del__(self):
        ''' If you forgot to save, this saves when the instance dies '''
        if self.wb:
            self.save()
        logging.debug("%s destructor: %s" % (self.__class__, self.wbName))

    # --------------------------------------------------------------------
    def ts2tabName(self, posixTimeStamp):
        ''' Tab names are: '2016-05-01' '''
        # WARNING - changing this format will change the way tabs are sorted!
        tabDate = datetime.date.fromtimestamp(posixTimeStamp)
        tabName = "%04d-%02d-%02d" % (tabDate.year, tabDate.month, tabDate.day)
        return tabName

    # --------------------------------------------------------------------
    def foreach_cell(self, ws, cell_range, **kwargs):
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                for attrName in kwargs:
                    setattr(cell, attrName, kwargs[attrName])


    # --------------------------------------------------------------------
    def subtotals(self, ws, startCols):

        # -------------
        def subtotal_helper(sumRow, sumCols, sumIdx, col1Idx, colNIdx):
            ''' The formula for the subtotal will go in cell [sumCol, sumRow] '''
            sumCol = sumCols[sumIdx][col1Idx]   # Column is always col1Idx
            # Recursion logic:
            # For this level, the sum is in columns sumCols[sumIdx][col1Idx:colNIdx]
            # To find col1Idx and colNIdx for the recursion, we find the indexes of
            # each pair of letters in this level's sumCols list in the next level's
            # sumCol list.  The range for the recursion is this interval minus 1.
            #
            for i in range(col1Idx, colNIdx):
                colStr1 = sumCols[sumIdx][i]        # First letter pair
                colStrN = sumCols[sumIdx][i+1]      # Next letter pair
                nxtcol1Idx = sumCols[sumIdx + 1].index(colStr1)     # Index @next
                nxtcolNIdx = sumCols[sumIdx + 1].index(colStrN)     # Index @next
                subtotal_helper(sumRow-1, sumCols, sumIdx+1, nxtcol1Idx, nxtcolNIdx)

            # Now insert sum at this level
            sumCells = [[], []]            # Build Total/Fail lists

            # On the last subtotal, we sum the test columns, otherwise we sum sums
            if sumIdx == len(sumCols) - 1:
                row1Idx = self.ROW_TESTS
                rowNIdx = row1Idx + len(self.TEST_NAMES)
                testCol = sumCol
                for j in range(2):
                    for row in range(row1Idx, rowNIdx):
                        cellAddr = "{}{}".format(testCol, row)
                        sumCells[j].append(cellAddr)
                    testCol = self.ccNext(testCol)
            else:
                # Collect the cells for this level
                for i in range(col1Idx, colNIdx):
                    str_col = sumCols[sumIdx][i]
                    for j in range(2):
                        cellAddr = "{}{}".format(str_col, sumRow - 1)
                        sumCells[j].append(cellAddr)
                        str_col = self.ccNext(str_col)

            # --- Total/Fail cells
            for i in range(2):
                sumAddr = "{}{}".format(sumCol, sumRow)
                sum = "=SUM(%s)" % (','.join(sumCells[i]))
                ws[sumAddr].value = sum
                ws[sumAddr].font = Font(color=XLSX.cs(cFont))
                ws[sumAddr].fill = self.solidFill(cFill)
                sumCol = self.ccNext(sumCol)

            # --- '% Pass' cell
            cellAddr = "{}{}".format(sumCol, sumRow)
            cell_range = "{0}:{0}".format(cellAddr)
            self.foreach_cell(ws, cell_range,
                              value=XLSX.RESULT_FORMULA,
                              number_format=XLSX.RESULT_FORMAT,
                              font=Font(color=XLSX.cs(cFont)),
                              fill=self.solidFill(cFill)
            )
            ws.conditional_formatting.add(cell_range, XLSX.RESULT_RULE)
# startCols = {
#     'OS': ['C', 'K', 'W', 'AI'],
#     'BROWSER': ['C', 'G', 'K', 'O', 'S', 'W', 'AA', 'AE', 'AI'],
#     'RESULT': ['C', 'D', 'E', 'G', 'H', 'I', 'K', 'L', 'M', 'O', 'P', 'Q', 'S', 'T', 'U', 'W', 'X', 'Y', 'AA', 'AB', 'AC', 'AE', 'AF', 'AG', 'AI', 'AJ', 'AK'],
# }

        # -------------
        sumCols = []
        sumCols.append(startCols[OS])
        sumCols.append(startCols[BROWSER])
        sumCols.append(startCols[RESULT])
        subtotal_helper(14, sumCols, 0, 0, len(startCols[OS])-1)

    # --------------------------------------------------------------------
    def createTab(self, tabName):
        ''' Initialize a new tab in the workbook '''
        if (self.wb and tabName in self.wb.sheetnames):
            logging.info("Tab \'%s\' is exists..." % (tabName))
            self.setActiveByName(tabName)
            return self.wb[tabName]

        logging.debug("Tab \'%s\' is being created..." % (tabName))

        # Insert this sheet in the right place (keep tabs sorted)
        sheetNames = list(self.wb.sheetnames)
        logging.debug("WB sheets: %s" % (sheetNames))
        sheetNames.append(tabName)
        sheetNames.sort()
        sheetNames.reverse()
        logging.debug("WB.rev(): %s" % (sheetNames))
        i = sheetNames.index(tabName)
        ws = self.wb.create_sheet(tabName, i)
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        self.setActiveByName(tabName)
        logging.debug("WB final: %s" % (self.wb.sheetnames))

        # --- Populate the new tab framework ---
        ws['A1'] = "Created tab:"
        ws['A2'] = str(datetime.datetime.now()).split('.')[0][2:]

        # Set up ID and Test Name
        baseRow = self.ROW_TITLES
        cellID = ws['A' + str(baseRow)]
        cellTest = ws['B' + str(baseRow)]
        cellID.value = 'ID'
        cellID.font = Font(color=XLSX.cs(self.COLOR_ID_FONT))
        cellID.fill = self.solidFill(self.COLOR_ID_BKGD)
        cellTest.value = 'Test Name'
        cellTest.font = Font(color=XLSX.cs(self.COLOR_TEST_FONT))
        cellTest.fill = self.solidFill(self.COLOR_TEST_BKGD)
        for i, testName in enumerate(self.TEST_NAMES):
            ws['A' + str(self.ROW_TESTS + i)] = i + 1
            ws['B' + str(self.ROW_TESTS + i)] = testName
        self.wb.active.column_dimensions['A'].width = self.WIDTH_ID
        self.wb.active.column_dimensions['B'].width = self.WIDTH_TEST
        rt = self.ROW_TITLES
        cell_range = "A{}:B{}".format(rt, rt + len(self.TEST_NAMES))
        self.foreach_cell(ws, cell_range, border=XLSX.THIN_BORDER)

        # Set up OS/Browser/Total-Passed-Failed
        startCols = {OS: [], BROWSER: [], RESULT: []}
        alignHC = Alignment(horizontal="center")
        osCol = 'C'
        row = self.ROW_OS
        # --- OS LOOP ---
        for layoutIdx in range(len(self.LAYOUT)):
            startCols[OS].append(osCol)
            (osName, broswerList) = self.layout(layoutIdx)
            cell = ws[osCol+str(row)]
            cell.value = osName
            cell.fill = self.solidFill(self.COLOR_OS[osName])
            cell.alignment = alignHC
            lastID = ((layoutIdx + 1) == len(self.LAYOUT))

            broswerCol = osCol
            lastBrowser = False
            # --- BROWSER LOOP ---
            for browserName in broswerList:
                startCols[BROWSER].append(broswerCol)
                lastBrowser = (broswerList[-1] == browserName)
                cFont = self.COLOR_BROWSER[browserName]['font']
                cFill = self.COLOR_BROWSER[browserName]['fill']
                cell = ws[broswerCol+str(self.ROW_BROWSER)]
                cell.value = browserName
                cell.font = Font(color=XLSX.cs(cFont))
                cell.fill = self.solidFill(cFill)
                cell.alignment = alignHC

                resultCol = broswerCol
                # --- RESULT LOOP ---
                for resultName in self.RESULT_NAMES:
                    startCols[RESULT].append(resultCol)
                    cell = ws[resultCol+str(self.ROW_TITLES)]
                    cell.value = resultName
                    cell.fill = self.solidFill(self.COLOR_RESULT[resultName])
                    cell.alignment = alignHC
                    resultCol = XLSX.ccNext(resultCol)
                # --- RESULT LOOP END ---

                endCol = XLSX.ccPrev(resultCol)
                merge = "{0}{1}:{2}{1}".format(broswerCol, str(row+1),
                                               XLSX.ccPrev(resultCol))
                ws.merge_cells(merge)
                rt = self.ROW_TITLES
                cell_range = "{}{}:{}{}".format(
                    broswerCol, rt,
                    endCol, rt + len(self.TEST_NAMES)
                )
                self.foreach_cell(ws, cell_range, border=XLSX.THIN_BORDER)
                cell_range = "{}{}:{}{}".format(
                    endCol, rt+1,
                    endCol, rt + len(self.TEST_NAMES) + 1   # Plus 1 for subtot
                )
                self.foreach_cell(ws, cell_range,
                                  value=XLSX.RESULT_FORMULA,
                                  number_format=XLSX.RESULT_FORMAT
                                  )
                ws.conditional_formatting.add(cell_range, XLSX.RESULT_RULE)
                broswerCol = resultCol  # Start next browser at last result col
                if not lastBrowser:   # Leave a thin col after except for last
                    ws.column_dimensions[broswerCol].width = \
                        self.WIDTH_OS_GAP / 2
                    broswerCol = XLSX.ccNext(broswerCol)
            # --- BROWSER LOOP END ---

            merge = "{0}{1}:{2}{1}".format(osCol, row, XLSX.ccPrev(broswerCol))
            ws.merge_cells(merge)
            osCol = broswerCol  # Start next OS after last browser col
            if not lastID:   # Leave a skinny column after except for last
                ws.column_dimensions[osCol].width = self.WIDTH_OS_GAP
                osCol = XLSX.ccNext(osCol)
        # --- OS LOOP END ---

        # --- Insert TEST subtotals ---
        for testIdx in range(len(self.TEST_NAMES)+1):
            row = self.ROW_TESTS + testIdx
            cellList = [[],[]]
            for colIdx in range(len(startCols[BROWSER])-1):
                colC = startCols[BROWSER][colIdx]
                cellList[0].append("{}{}".format(colC, row))
                cellList[1].append("{}{}".format(self.ccNext(colC), row))
            str_col = startCols[BROWSER][-1]
            for i in range(2):
                sumFormula = "=SUM(%s)" % (','.join(cellList[i]))
                cellAddr = "{}{}".format(str_col, row)
                ws[cellAddr].value = sumFormula
                str_col = self.ccNext(str_col)

        sumCols = []
        for i in [OS, BROWSER, RESULT]:
            sumCols.append(startCols[i])

        sumColors = []
        for i in [GRAND, OS, BROWSER]:
            t = (self.COLOR_SUBTOTAL[i]['font'], self.COLOR_SUBTOTAL[i]['fill'])
            sumColors.append(t)

        # --- Subtotal recursion helper ---
        def subtotal_helper(sumRow, sumIdx, col1Idx, colNIdx):
            ''' The formula for the subtotal will go in cell [sumCol, sumRow] '''

            sumCol = sumCols[sumIdx][col1Idx]   # Column is always col1Idx
            # Recursion logic:
            # For this level, the sum is in columns sumCols[sumIdx][col1Idx:colNIdx]
            # To find col1Idx and colNIdx for the recursion, we find the indexes of
            # each pair of letters in this level's sumCols list in the next level's
            # sumCol list.  The range for the recursion is this interval minus 1.
            #
            if sumIdx < len(sumCols)-1:
                for i in range(col1Idx, colNIdx):
                    colStr1 = sumCols[sumIdx][i]        # First letter pair
                    colStrN = sumCols[sumIdx][i+1]      # Next letter pair
                    nxtcol1Idx = sumCols[sumIdx + 1].index(colStr1)     # Index @next
                    nxtcolNIdx = sumCols[sumIdx + 1].index(colStrN)     # Index @next
                    subtotal_helper(sumRow-1, sumIdx+1, nxtcol1Idx, nxtcolNIdx)

            # Now insert sum at this level
            sumCells = [[], []]            # Build Total/Fail lists

            # On the last subtotal, we sum the test columns, otherwise we sum sums
            if sumIdx == len(sumCols) - 1:
                row1Idx = self.ROW_TESTS
                rowNIdx = row1Idx + len(self.TEST_NAMES)
                testCol = sumCol
                for i in range(2):
                    cellAddr = "{0}{1}:{0}{2}".format(testCol, row1Idx, rowNIdx-1)
                    sumCells[i].append(cellAddr)
                    testCol = self.ccNext(testCol)
            else:
                # Collect the cells for this level
                for i in range(col1Idx, colNIdx):
                    str_col = sumCols[sumIdx][i]
                    for j in range(2):
                        cellAddr = "{}{}".format(str_col, sumRow - 1)
                        sumCells[j].append(cellAddr)
                        str_col = self.ccNext(str_col)

            # --- Fill in the sums for the Total & Fail cells
            (cFont, cFill) = sumColors[sumIdx]
            for i in range(2):
                sumAddr = "{}{}".format(sumCol, sumRow)
                sum = "=SUM(%s)" % (','.join(sumCells[i]))
                ws[sumAddr].value = sum
                ws[sumAddr].font = Font(color=XLSX.cs(cFont))
                ws[sumAddr].fill = self.solidFill(cFill)
                sumCol = self.ccNext(sumCol)

            # --- '% Pass' cell
            cellAddr = "{}{}".format(sumCol, sumRow)
            cell_range = "{0}:{0}".format(cellAddr)
            self.foreach_cell(ws, cell_range,
                              value=XLSX.RESULT_FORMULA,
                              number_format=XLSX.RESULT_FORMAT,
                              font=Font(color=XLSX.cs(cFont)),
                              fill=self.solidFill(cFill)
            )
            ws.conditional_formatting.add(cell_range, XLSX.RESULT_RULE)
# startCols = {
#     'OS': ['C', 'K', 'W', 'AI'],
#     'BROWSER': ['C', 'G', 'K', 'O', 'S', 'W', 'AA', 'AE', 'AI'],
#     'RESULT': ['C', 'D', 'E', 'G', 'H', 'I', 'K', 'L', 'M', 'O', 'P', 'Q', 'S', 'T', 'U', 'W', 'X', 'Y', 'AA', 'AB', 'AC', 'AE', 'AF', 'AG', 'AI', 'AJ', 'AK'],
# }
        # --- Prime the subtotal recursion ---
        subtotal_helper(14, 0, 0, len(startCols[OS])-1)

        # --- Prime the subtotal recursion ---
        baseRow = self.ROW_TITLES + len(self.TEST_NAMES) + 1
        for i in range(len(self.SUBTOTAL_NAMES)):
            cell = ws['B' + str(baseRow + i)]
            cell.value = self.SUBTOTAL_NAMES[i]
            cell.alignment = Alignment(horizontal="right")


# ------------------------------------------------------------------------------
class Logfile(object):
    def __init__(self, fname):
        self.__time_init = time.time()  # Time to init this classs
        self.__fname = fname            # File name of log
        self.__f_dt = datetime.datetime.fromtimestamp(os.stat(fname).st_mtime)
        self.__tname = None             # Test name from log
        self.__tname_os = None          # OS name
        self.__tname_browser = None     # Browser name
        self.__tname_suite = None       # Test suite name
        self.__log = None               # Content of log file with bad ASCII out
        self.__time_us = None           # Processing time in µs
        self.__t_fails = {}             # Maps failure name to count
        self.__t_pends = {}             # Maps pending test name to count
        self.__tests = 0                # Tests run
        self.__fails = 0                # Tests failed
        self.__pends = 0             # Tests skipped
        self.__log_bundle_fname = None
        self.__load_file()              # Load the data
        self.__time_init = time.time() - self.__time_init
        logging.info("\t%-10s: %f (ms)" % ("Class init time", self.__time_init * 1000))

    @property
    def file_name(self):
        return self.__fname

    @property
    def file_datetime(self):
        return self.__f_dt

    @property
    def test_name(self):
        return self.__tname

    @property
    def test_name_os(self):
        return self.__tname_os

    @property
    def test_name_browser(self):
        return self.__tname_browser

    @property
    def test_name_suite(self):
        return self.__tname_suite

    @property
    def num_tests(self):
        return self.__tests

    @property
    def num_fails(self):
        return self.__fails

    @property
    def num_pending(self):
        return self.__pends

    @property
    def tests_failed_map(self):
        return self.__t_fails

    @property
    def tests_pending_map(self):
        return self.__t_pends

    @property
    def log(self):
        return self.__log

    @property
    def time_fread(self):
        return self.__time_us

    @property
    def time_init(self):
        return self.__time_init

    @property
    def log_bundle_fname(self):
        return self.__log_bundle_fname

    TEST_MATRIX = {
      XLSX.TEST_NAMES[0]: [XLSX.OS_UBUNTU, XLSX.OS_MACOSX], # VC60
      XLSX.TEST_NAMES[1]: [XLSX.OS_UBUNTU, XLSX.OS_MACOSX], # VC65
      XLSX.TEST_NAMES[2]: [XLSX.OS_UBUNTU, XLSX.OS_MACOSX], # VC60
      XLSX.TEST_NAMES[3]: [XLSX.OS_UBUNTU, XLSX.OS_MACOSX], # VC65
      XLSX.TEST_NAMES[4]: [ # VC60
        {XLSX.OS_UBUNTU: [XLSX.BROWSER_CHROME, XLSX.BROWSER_FIREFOX]},
        {XLSX.OS_MACOSX: [XLSX.BROWSER_CHROME, XLSX.BROWSER_FIREFOX, XLSX.BROWSER_SAFARI]},
        {XLSX.OS_WINDOWS: [XLSX.BROWSER_CHROME, XLSX.BROWSER_FIREFOX]}
      ],
      XLSX.TEST_NAMES[5]: [ # VC65
        {XLSX.OS_UBUNTU: [XLSX.BROWSER_CHROME, XLSX.BROWSER_FIREFOX]},
        {XLSX.OS_MACOSX: [XLSX.BROWSER_CHROME, XLSX.BROWSER_FIREFOX, XLSX.BROWSER_SAFARI]},
        {XLSX.OS_WINDOWS: [XLSX.BROWSER_CHROME, XLSX.BROWSER_FIREFOX, XLSX.BROWSER_IE]}
      ]
    }
    # --------------------------------------------------------------------
    def __del__(self):
        ''' If you forgot to save, this saves when the instance dies '''
        logging.debug("%s destructor..." % (self.__class__))


    def __load_file(self):
        '''
        Return the Specs, Fails, and File Mod Date/Time
        '''
        # Process the file, but catch any errors reading it in...
        logging.info("Processing file: \'%s\'" % (self.__fname))
        t1 = time.time()
        try:
            # We open the file in binary mode to handle the screen color escape
            # codes that get interpreted as bad Unicode characters.
            # Then the .decode() cleans up the bytes by removing any high ASCII
            # values, which is all we need for this processing.
            with open(self.__fname, 'rb') as f:
                self.__log = f.read().decode('ascii', 'replace')
        except Exception as e:
            raise e("Read/Decode error in file: \'%s\'" % (self.__fname))
        self.__time_us = time.time() - t1

        # Find the test name or die
        test_tree = ET.parse(self.__fname)
        self.__tname = "-"

        def set_os_and_browser():
            test_dirname = os.path.dirname(self.__fname).split('/')
            test_dirname = test_dirname[len(test_dirname) - 1].split('-')
            test_dirname_len = len(test_dirname)
            is_browser_info_needed = True
            if test_dirname[1] == "4":
                self.__tname_browser = normalize_browsername(test_dirname[test_dirname_len - 1])
                self.__tname_os = test_dirname[test_dirname_len - 2]
            else:
                # defaulting to chrome, as 18-1, 18-2 and 18-3 do not require browser
                self.__tname_browser = XLSX.BROWSER_CHROME
                self.__tname_os = test_dirname[test_dirname_len - 1]
                is_browser_info_needed = False
            return is_browser_info_needed

        def normalize_browsername(browser):
            m = RE_BROWSER_CHROME.search(browser)
            if m:
                normalized = browser[m.start():m.end()]
                return normalized[0].upper() + normalized[1:]
            m = RE_BROWSER_IE.search(browser)
            if m:
                return XLSX.BROWSER_IE

            return browser[0].upper() + browser[1:]

        def get_log_bundle_name():
            date_string = self.__f_dt.strftime("%m%d%y")
            os_string = "%s" % (self.__tname_os)
            browser_string = "%s" % ("-" + self.__tname_browser if is_browser_info_needed else "")
            return "%s-%s%s/log.html" % (self.__tname_suite, os_string, browser_string)

        is_browser_info_needed = set_os_and_browser()
        self.__tname_suite = "%s-%s" % (test_tree.find("suite").attrib["name"], "60" if "-60-" in self.__fname else "65")  # Test suite name
        self.__log_bundle_fname = get_log_bundle_name()

        stats_all_element = test_tree.find("statistics/total/stat[2]")
        self.__tests = 0 if is_browser_info_needed else (int(stats_all_element.attrib["fail"]) + int(stats_all_element.attrib["pass"]))
        self.__fails = 0 if is_browser_info_needed else int(stats_all_element.attrib["fail"])

        bucket = self.__t_fails

        # if the current suite is 18-4 extract details from from ngc_tests.log
        if is_browser_info_needed:
            log_filename = "ui-test-results/%s-%s-%s/ngc_tests.log" % (self.__tname_suite, self.__tname_os, self.__tname_browser)
            try:
                with open(log_filename, 'rb') as f:
                    log = f.read().decode('ascii', 'replace')
                    suite_failed = False
                    if "-65-" in self.__fname:
                        for line in log.split("\n"):
                            if RE_VIC_H5_TEST_FAIL_START.match(line):
                                suite_failed = True
                                data = "%s (%s - see ngc_tests.log)" % (self.__tname_suite, RE_VIC_H5_TEST_FAIL_START.match(line).groups()[0])
                                bucket.setdefault(data, 0)
                                bucket[data] += 1
                                self.__fails += 1
                                self.__tests += 1
                            elif RE_VIC_H5_TEST_PASS_START.match(line):
                                self.__tests += 1

                    else:
                        found_failed_test = False
                        for line in log.split("\n"):
                            if found_failed_test:
                                data = "%s (%s - see ngc_tests.log)" % (self.__tname_suite, line.strip())
                                bucket.setdefault(data, 0)
                                bucket[data] += 1
                                found_failed_test = False
                                continue
                            if RE_VIC_FLEX_TEST_FAIL_START.match(line):
                                found_failed_test = True
                                suite_failed = True
                                self.__fails += 1
                                self.__tests += 1
                                continue
                            elif RE_VIC_FLEX_TEST_PASS_START.match(line):
                                self.__tests += 1
            except FileNotFoundError as e:
                logging.warning("ngc_tests.log was not found for %s" % self.__tname_suite)

            except Exception as e:
                raise e("Read/Decode error in file: \'%s\'" % (log_filename))

        for test in test_tree.find("suite").iter("test"):
            status_element = test.find("status")
            test_name = test.attrib["name"]
            if test_name == "Run Ngc Tests Project":
                continue

            if status_element.attrib["status"] == "FAIL":
                data = "%s (%s)" % (self.__tname_suite, test_name)
                bucket.setdefault(data, 0)
                bucket[data] += 1

        # Return the relevant data for our log file
        FMT = "\t%-20s: %s"
        logging.info(FMT % ("File name", self.file_name))
        logging.info(FMT % ("File Mod Time", self.file_datetime))
        logging.info(FMT % ("File read time (ms)", str(self.time_fread * 1000)))
        logging.info(FMT % ("Test Name", self.test_name))
        logging.info(FMT % ("OS Name", self.test_name_os))
        logging.info(FMT % ("Browser", self.test_name_browser))
        logging.info(FMT % ("Suite Name", self.test_name_suite))
        logging.info(FMT % ("Tests RUN", self.num_tests))
        logging.info(FMT % ("Tests FAILED", self.num_fails))
        logging.info(FMT % ("Pending tests", self.num_pending))

# ------------------------------------------------------------------------------
def main(args):
    ''' --- Main routine --- '''

    # Get the real names of the logging levels
    LOG_LEVEL_NAMES = []
    for _ in [logging.DEBUG, logging.INFO, logging.WARNING, logging.ERROR]:
        LOG_LEVEL_NAMES.append(logging.getLevelName(_))

    # Parse the arguments
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "DIR",
        default="/var/lib/jenkins/jobs",
        metavar = "searchdir",
        help="Search this directory for log files to parse"
                        " (default=\'%(default)s\')"
    )
    parser.add_argument(
        "-s", "--skipdir",
        default=["analysis", "demo", "deploy"],
        nargs='+',
        metavar = "str",
        help="Directory paths with this in their names will be skipped"
            " (default=\'%(default)s\')"
    )
    parser.add_argument(
        "-x", "--excel",
        default="log.xlsx",
        metavar = "log.xlsx",
        dest="xlsx",
        help="Specify .xlsx file to update (or create)"
            " (default=\'%(default)s\')"
    )
    parser.add_argument(
        "-f", "--logfile",
        default="log",
        dest="logfile",
        metavar = "log",
        help="Name of log files (default=\'%(default)s\')"
    )
    parser.add_argument(
        "-l", "--log",
        default="WARNING",
        choices=LOG_LEVEL_NAMES,
        dest="logLvl",
        help="Set the logging level (default=\'%(default)s\')"
    )
    parser.add_argument(
        "-d", "--debug",
        default=logging.WARNING,
        action="store_const", dest="loglevel", const=logging.DEBUG,
        help="Enable debugging statements (same as --log DEBUG)"
    )
    parser.add_argument(
        "-v", "--verbose",
        dest="loglevel", action="store_const",  const=logging.INFO,
        help="Be verbose (same as --log INFO)"
    )

    args = parser.parse_args(args)
    assert args.DIR is not None, "--dir: Directory to search must be specified"
    assert args.xlsx is not None, "--excel: Excel file name must be specified"

    # Set up the logger
    lfmt = "%(asctime)s: [%(levelname)-7s: %(funcName)-10s#%(lineno)4d] %(message)s"
    ll = min(args.loglevel, getattr(logging, args.logLvl))
    logging.basicConfig(format=lfmt, datefmt='%y.%m.%d %H:%M:%S', level=ll)
    logging.addLevelName( logging.WARNING, "\033[1;33m%s\033[1;0m" % logging.getLevelName(logging.WARNING))
    logging.addLevelName( logging.ERROR, "\033[1;41m%s\033[1;0m" % logging.getLevelName(logging.ERROR))
    logging.debug("DEBUG: Arguments: %s" % (args))

    # Open the output XSLX file, creates it if it does not exist
    xlsx = XLSX(args.xlsx)

    # Process the log file(s).
    logsPassFail = [0, 0]
    logsBad = []

    for dirpath, dirnames, filenames in os.walk(args.DIR):
        # Make the name of a log file in this directory
        targetLog = os.path.join(dirpath, args.logfile)

        logging.debug("%s<%s>|%s<%s>|%s<%s>" %
                      ("target", targetLog,
                       "dirnames", dirnames,
                       "filenames", filenames)
                      )
        # For the directories in 'dirpath', remove any ones we should skip
        tmpdirnames = dirnames
        for skipname in args.skipdir:
            for dirname in tmpdirnames:
                if dirname.find(skipname) >= 0:
                    dirnames.remove(dirname)
        logging.debug("dirnames after: %s" % (dirnames))

        # If the log file isn't in this directory, skip it.
        if not(os.path.exists(targetLog) and os.path.isfile(targetLog)):
            logging.debug("No log file in this directory")
            continue

        #print("Processing file <{}>".format(targetLog))
        try:
            # Parse the log file
            logging.debug("-" * 79)
            logfile = Logfile(targetLog)
            logging.debug("-" * 79)
            dt = logfile.file_datetime
            if dt.year < 2016 or (dt.year == 2016 and dt.month < 5):
                logging.info("Skipping old file: %s" % (dt))
                raise ValueError("File too old to process: %s" % (repr(dt)))
        except ValueError as e:
            logsBad.append(targetLog)
            logsPassFail[1] += 1
            continue

        st = datetime.datetime.strftime(dt, '%Y-%m-%d %H:%M:%S')

        try:
            # Update the Excel file
            xlsx.add(st, logfile)
        except ValueError as e:
            msg = "Bad os/browser/test \'%s\': %s" % (targetLog, e)
            logging.error(msg)
            logsBad.append(targetLog)
            logsPassFail[1] += 1
            continue

        logsPassFail[0] += 1

    logging.debug("Log files pass/fail: {}".format(logsPassFail))

    if logsBad:
        logging.debug("Malformed file names:")
        for i, l in enumerate(sorted(logsBad)):
            logging.debug("  #{:6d}: {}".format(i, l))

    del(xlsx)
    logging.info("Processing complete, exiting...")
    return


# -----------------------------------------------------------------------------
#    MAIN CODE
#
if __name__ == "__main__":
    try:
        sys.exit(main(sys.argv[1:]))
    except KeyboardInterrupt:
        print >> sys.stderr, \
            os.path.splitext(os.path.basename(__file__))[0] + ': interrupted'
        sys.exit(_EXIT_CONTROL_C_EXCEPTION)
